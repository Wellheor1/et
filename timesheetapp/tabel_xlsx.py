import calendar
import datetime

import pytils as pytils
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle, Font, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.cell import get_column_letter

department_table_number = 27
month_name = pytils.dt.ru_strftime(u"%B", inflected=True, date=datetime.datetime.now())
current_year = datetime.date.today().year
current_month = datetime.date.today().month
first_day_month = 1
last_day_month = calendar.monthrange(current_year, current_month)[1]
tabel_type = 'первичный'
hospital_name = 'ОГАУЗ ГИМДКБ'
department_name = 'Кабинет неотложной травматологии и ортопедии (травмпункт)'
date_now = datetime.datetime.now().strftime('%d.%m.%Y')
main_doctor = 'Новожилов В.А.'
head_department = 'Преториус Т.Л.'
old_sestra = 'Тотьямина Д.С.'
hr_specialist = 'Краснова С.А.'
form_okyd = '0504421'

book = Workbook()
active_list = book.active
active_list.page_setup.orientation = active_list.ORIENTATION_LANDSCAPE
active_list.page_setup.paperSize = active_list.PAPERSIZE_A4
active_list.page_setup.fitToPage = 1
active_list.page_margins = PageMargins(top=0.2, bottom=0.2, left=0.2, right=0.2)
active_list.title = 'Табель 52Н'

styleCenterTitle = NamedStyle(name='styleCenterTitle')
styleCenterTitle.alignment = Alignment(horizontal='center', vertical='center')

styleCenterSup = NamedStyle(name='styleCenterSup')
styleCenterSup.font = Font(size=6)
styleCenterSup.alignment = Alignment(horizontal='center', vertical='top')

styleCenterBorder = NamedStyle(name='styleCenterTitleBorder')
bd = Side(style='thin', color='000000')
styleCenterBorder.border = Border(left=bd, right=bd, top=bd, bottom=bd)
styleCenterBorder.font = Font(size=8)
styleCenterBorder.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

styleCenterDataBorder = NamedStyle(name='styleCenterDataBorder')
bd1 = Side(style='thin', color='000000')
styleCenterDataBorder.border = Border(left=bd1, right=bd1, top=bd1, bottom=bd1)
styleCenterDataBorder.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

styleCenterBorderBold = NamedStyle(name='styleCenterDataBorderBold')
bd2 = Side(style='thin', color='000000')
styleCenterBorderBold.border = Border(left=bd2, right=bd2, top=bd2, bottom=bd2)
styleCenterBorderBold.font = Font(bold=True, size=10)
styleCenterBorderBold.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

styleCenterBorderBottom = NamedStyle(name='styleCenterBorderBottom')
bd = Side(style='thin', color='000000')
styleCenterBorderBottom.border = Border(bottom=bd)
styleCenterBorderBottom.alignment = Alignment(horizontal='center', vertical='center')

styleCenterBorderBottomMin = NamedStyle(name='styleCenterBorderBottomMin')
bd = Side(style='thin', color='000000')
styleCenterBorderBottomMin.border = Border(bottom=bd)
styleCenterBorderBottomMin.font = Font(size=8)
styleCenterBorderBottomMin.alignment = Alignment(horizontal='center', vertical='bottom')


styleCenterTitleBold = NamedStyle(name='styleCenterTitleBold')
styleCenterTitleBold.alignment = Alignment(horizontal='center', vertical='center')
styleCenterTitleBold.font = Font(bold=True, color='000000')

styleRightTitle = NamedStyle(name='styleRightTitle')
styleRightTitle.font = Font(size=8)
styleRightTitle.alignment = Alignment(horizontal='right', vertical='center')

styleLeftTitle = NamedStyle(name='styleLeftTitle')
styleLeftTitle.font = Font(size=8)
styleLeftTitle.alignment = Alignment(horizontal='left', vertical='center')

styleRightTitleBold = NamedStyle(name='styleRightTitleBold')
styleRightTitleBold.alignment = Alignment(horizontal='right', vertical='center')
styleRightTitleBold.font = Font(bold=True, color='000000', size=8)

active_list.merge_cells(start_row=1, start_column=34, end_row=1, end_column=38)
active_list['AH1'].value = 'Утв приказом Минфина России'
active_list['AH1'].style = styleRightTitleBold

active_list.merge_cells(start_row=2, start_column=35, end_row=2, end_column=38)
active_list['AI2'].value = 'от 30 марта 2015г № 52н'
active_list['AI2'].style = styleRightTitleBold

active_list.merge_cells(start_row=3, start_column=4, end_row=3, end_column=30)
active_list['D3'].value = f'Табель № {department_table_number}'
active_list['D3'].style = styleCenterTitleBold

active_list.merge_cells(start_row=4, start_column=4, end_row=4, end_column=30)
active_list['D4'].value = 'учета использования рабочего времени'
active_list['D4'].style = styleCenterTitle
active_list.merge_cells(start_row=4, start_column=37, end_row=4, end_column=38)
active_list['AK4'].value = 'коды'

active_list['AJ5'].value = 'форма ОКУД'
active_list['AJ5'].style = styleRightTitle
active_list.merge_cells(start_row=5, start_column=37, end_row=5, end_column=38)
active_list['AK5'].value = f'{form_okyd}'

active_list.merge_cells(start_row=6, start_column=4, end_row=6, end_column=30)
active_list['D6'].value = f'За период с {first_day_month} по {last_day_month} {month_name}'
active_list['D6'].style = styleCenterTitle
active_list['AJ6'].value = 'Дата'
active_list['AJ6'].style = styleRightTitle
active_list.merge_cells(start_row=6, start_column=37, end_row=6, end_column=38)
active_list['AK6'].value = f'{date_now}'

active_list['A7'].value = 'Учреждение'
active_list['A7'].style = styleLeftTitle
active_list.merge_cells(start_row=7, start_column=4, end_row=7, end_column=30)
active_list['D7'].value = f'{hospital_name}'
# active_list.merge_cells(start_row=7, start_column=35, end_row=7, end_column=36)
active_list['AJ7'].value = 'по ОКПО'
active_list['AJ7'].style = styleRightTitle
active_list.merge_cells(start_row=7, start_column=37, end_row=7, end_column=38)

active_list.merge_cells(start_row=8, start_column=1, end_row=8, end_column=3)
active_list['A8'].value = 'Структурное подразделение'
active_list['A8'].style = styleLeftTitle
active_list.merge_cells(start_row=8, start_column=4, end_row=8, end_column=30)
active_list['D8'].value = f'{department_name}'
active_list.merge_cells(start_row=8, start_column=37, end_row=8, end_column=38)

active_list['A9'].value = 'Вид табеля'
active_list['A9'].style = styleLeftTitle
active_list.merge_cells(start_row=9, start_column=4, end_row=9, end_column=30)
active_list['D9'].value = f'{tabel_type}'
active_list.merge_cells(start_row=9, start_column=33, end_row=9, end_column=36)
active_list['AG9'].value = 'Номер корректировки'
active_list['AG9'].style = styleRightTitle
active_list.merge_cells(start_row=9, start_column=37, end_row=9, end_column=38)

active_list.merge_cells(start_row=10, start_column=4, end_row=10, end_column=30)
active_list['D10'].value = '(первичный - 0; корректирующий 1,2 и т.д.)'
active_list.merge_cells(start_row=10, start_column=32, end_row=10, end_column=36)
active_list['AF10'].value = 'Дата формирования документа'
active_list['AF10'].style = styleRightTitle
active_list.merge_cells(start_row=10, start_column=37, end_row=10, end_column=38)
active_list['AK10'].value = f'{date_now}'

for column in active_list['AK4':'AL10']:
    for cell in column:
        cell.style = styleCenterBorder

for column in active_list['D7':'AD9']:
    for cell in column:
        cell.style = styleCenterBorderBottom
    active_list['D8'].font = Font(bold=True)

for column in active_list['D10':'AD10']:
    for cell in column:
        cell.style = styleCenterSup

column_name = get_column_letter(last_day_month+9)

active_list.merge_cells(start_row=12, start_column=4, end_row=12, end_column=last_day_month+9)
active_list['D12'].value = 'Числа месяца'
for column in active_list['D12':f'{column_name}12']:
    for cell in column:
        cell.style = styleCenterDataBorder


for column in active_list['D13': f'{column_name}13']:
    counter = 1
    for cell in column:
        if counter < 16:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 3.5
            cell.value = counter
        elif counter == 16:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 7
            cell.value = 'Итого дней (часов) явок (неявок) с 1-15'
        elif counter < last_day_month+2:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 3.5
            cell.value = counter-1
        elif counter == last_day_month+2:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 7
            cell.value = 'Всего дней (часов) явок (неявок) за месяц'
        elif counter == last_day_month+3:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 5
            cell.value = 'Всего отработано часов'
        elif counter == last_day_month+4:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 5
            cell.value = 'Ночные'
        elif counter == last_day_month+5:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 5
            cell.value = 'Выходные'
        elif counter == last_day_month+6:
            active_list.column_dimensions[f'{get_column_letter(counter + 3)}'].width = 5
            cell.value = 'Праздничные'
        cell.style = styleCenterDataBorder
        cell.font = Font(size=8)
        counter += 1

active_list.column_dimensions['A'].width = 13
active_list.column_dimensions['B'].width = 4.7
active_list.column_dimensions['C'].width = 10
active_list.merge_cells(start_row=12, start_column=1, end_row=13, end_column=1)
active_list['A12'].value = 'Фамилия, имя, отчество'
active_list.merge_cells(start_row=12, start_column=2, end_row=13, end_column=2)
active_list['B12'].value = 'учетный номер'
active_list.merge_cells(start_row=12, start_column=3, end_row=13, end_column=3)
active_list['C12'].value = 'Должность (профессия)'

for column in active_list['A12':'C13']:
    for cell in column:
        cell.style = styleCenterDataBorder
        cell.font = Font(size=8)

employees = [
        {"person": "Прет111ориус Т.Л.", "tabel_number": "885", "post": "Заведующий кабинетом врач травматолог-ортопед",
         "hours15": [f'{x}' for x in range(1, 16)], "summ_15": '324', "hours30": [f'{x}' for x in range(16, last_day_month+1)],
         "common_days": "31", "common_hours": "555", "night_hours": '33', "weekends_hours": "30",
         "holidays_hours": "10"},
        {"person": "Прет2222ориус Т.Л.", "tabel_number": "8835", "post": "врач травматолог-ортопед внутр. совместитель",
         "hours15": [f'{x}' for x in range(1, 16)], "summ_15": '324', "hours30": [f'{x}' for x in range(16, last_day_month+1)],
         "common_days": "31", "common_hours": "555", "night_hours": '33', "weekends_hours": "30",
         "holidays_hours": "10"}
    ]
num_row = 13
for row, entry in enumerate(employees, start=1):
    active_list.cell(row=row + num_row, column=1, value=entry["person"])
    active_list.cell(row=row + num_row, column=1).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=2, value=entry["tabel_number"])
    active_list.cell(row=row + num_row, column=2).style = styleCenterBorderBold
    active_list.cell(row=row + num_row, column=3, value=entry["post"])
    active_list.cell(row=row + num_row, column=3).style = styleCenterBorder
    for idx, data in enumerate(entry["hours15"], 4):
        active_list.cell(row=row + num_row, column=idx, value=data)
        active_list.cell(row=row + num_row, column=idx).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=19, value=entry["summ_15"])
    active_list.cell(row=row + num_row, column=19).style = styleCenterDataBorder
    for idx, data in enumerate(entry["hours30"], 20):
        active_list.cell(row=row + num_row, column=idx, value=data)
        active_list.cell(row=row + num_row, column=idx).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=last_day_month + 5, value=entry["common_days"])
    active_list.cell(row=row + num_row, column=last_day_month + 5).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=last_day_month + 6, value=entry["common_hours"])
    active_list.cell(row=row + num_row, column=last_day_month + 6).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=last_day_month + 7, value=entry["night_hours"])
    active_list.cell(row=row + num_row, column=last_day_month + 7).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=last_day_month + 8, value=entry["weekends_hours"])
    active_list.cell(row=row + num_row, column=last_day_month + 8).style = styleCenterDataBorder
    active_list.cell(row=row + num_row, column=last_day_month + 9, value=entry["holidays_hours"])
    active_list.cell(row=row + num_row, column=last_day_month + 9).style = styleCenterDataBorder

max_row = active_list.max_row

active_list[f'A{max_row+2}'].value = 'Главный врач'
active_list[f'A{max_row+2}'].style = styleCenterBorderBottomMin
active_list.merge_cells(start_row=max_row+2, start_column=3, end_row=max_row+2, end_column=9)
active_list[f'C{max_row+2}'].value = f'{main_doctor}'
active_list[f'C{max_row+2}'].style = styleCenterBorderBottomMin
for column in active_list[f'C{max_row+2}':f'H{max_row+2}']:
    for cell in column:
        cell.style = styleCenterBorderBottomMin
        cell.font = Font(size=8)

active_list[f'A{max_row+4}'].value = 'Зав. отделения'
active_list[f'A{max_row+4}'].style = styleCenterBorderBottomMin
active_list.merge_cells(start_row=max_row+4, start_column=3, end_row=max_row+4, end_column=9)
active_list[f'C{max_row+4}'].value = f'{head_department}'
active_list[f'C{max_row+4}'].style = styleCenterBorderBottomMin
for column in active_list[f'C{max_row+4}':f'H{max_row+4}']:
    for cell in column:
        cell.style = styleCenterBorderBottomMin
        cell.font = Font(size=8)

active_list[f'A{max_row+6}'].value = 'Старшая медсестра'
active_list[f'A{max_row+6}'].style = styleCenterBorderBottomMin
active_list.merge_cells(start_row=max_row+6, start_column=3, end_row=max_row+6, end_column=9)
active_list[f'C{max_row+6}'].value = f'{old_sestra}'
active_list[f'C{max_row+6}'].style = styleCenterBorderBottomMin
for column in active_list[f'C{max_row+6}':f'H{max_row+6}']:
    for cell in column:
        cell.style = styleCenterBorderBottomMin

active_list[f'A{max_row+8}'].value = 'Специалист о.к.'
active_list[f'A{max_row+8}'].style = styleCenterBorderBottomMin
active_list.merge_cells(start_row=max_row+8, start_column=3, end_row=max_row+8, end_column=9)
active_list[f'C{max_row+8}'].value = f'{hr_specialist}'
active_list[f'C{max_row+8}'].style = styleCenterBorderBottomMin
for column in active_list[f'C{max_row+8}':f'H{max_row+8}']:
    for cell in column:
        cell.style = styleCenterBorderBottomMin


book.save('tabel.xlsx')